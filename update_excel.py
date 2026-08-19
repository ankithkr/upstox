import datetime
import json
import os

import requests
from openpyxl import load_workbook

from instruments import resolve_instrument_key

TOKEN_FILE = "token.json"
WORKBOOK_FILE = "watchlist.xlsx"
LTP_URL = "https://api.upstox.com/v3/market-quote/ltp"
HEADER_ROW = 4
FIRST_DATA_ROW = HEADER_ROW + 1
BATCH_SIZE = 50


def load_access_token():
    if not os.path.exists(TOKEN_FILE):
        raise SystemExit("No access token found - log in via /login first")
    with open(TOKEN_FILE) as f:
        return json.load(f)["access_token"]


def fetch_quotes(instrument_keys, access_token):
    quotes = {}
    for i in range(0, len(instrument_keys), BATCH_SIZE):
        batch = instrument_keys[i : i + BATCH_SIZE]
        response = requests.get(
            LTP_URL,
            params={"instrument_key": ",".join(batch)},
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Authorization": f"Bearer {access_token}",
            },
        )
        response.raise_for_status()
        for entry in response.json()["data"].values():
            quotes[entry["instrument_token"]] = entry
    return quotes


def main():
    access_token = load_access_token()

    wb = load_workbook(WORKBOOK_FILE)
    ws = wb.active

    rows = []
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        symbol = ws.cell(row=row, column=1).value
        if not symbol:
            continue
        symbol = str(symbol).strip().upper()
        try:
            instrument_key = resolve_instrument_key(symbol)
        except ValueError:
            ws.cell(row=row, column=2).value = "NOT FOUND"
            continue
        ws.cell(row=row, column=2).value = instrument_key
        rows.append((row, instrument_key))

    quotes = fetch_quotes([key for _, key in rows], access_token)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row, instrument_key in rows:
        quote = quotes.get(instrument_key)
        if not quote:
            ws.cell(row=row, column=3).value = "NO DATA"
            continue
        ws.cell(row=row, column=3).value = quote["last_price"]
        ws.cell(row=row, column=4).value = quote["ltq"]
        ws.cell(row=row, column=5).value = quote["volume"]
        ws.cell(row=row, column=6).value = quote["cp"]
        ws.cell(row=row, column=9).value = now

    wb.save(WORKBOOK_FILE)
    print(f"Updated {len(rows)} symbol(s) in {WORKBOOK_FILE}")


if __name__ == "__main__":
    main()

import json
import os
import time

from openpyxl import load_workbook
from openpyxl.styles import Font

from moving_average import fetch_20_day_average

TOKEN_FILE = "token.json"
WORKBOOK_FILE = "watchlist.xlsx"
FIRST_DATA_ROW = 5
DATA_FONT = Font(name="Arial", size=10)
CURRENCY_FORMAT = "₹#,##0.00;-₹#,##0.00;-"
PERCENT_FORMAT = "0.00%;(0.00%);-"


def load_access_token():
    if not os.path.exists(TOKEN_FILE):
        raise SystemExit("No access token found - log in via /login first")
    with open(TOKEN_FILE) as f:
        return json.load(f)["access_token"]


def main():
    access_token = load_access_token()

    wb = load_workbook(WORKBOOK_FILE)
    watchlist = wb["Watchlist"]
    analysis = wb["Analysis"]

    updated = 0
    for row in range(FIRST_DATA_ROW, watchlist.max_row + 1):
        symbol = watchlist.cell(row=row, column=1).value
        instrument_key = watchlist.cell(row=row, column=2).value
        if not symbol or not instrument_key or instrument_key == "NOT FOUND":
            continue

        avg = fetch_20_day_average(instrument_key, access_token)

        ma_cell = analysis.cell(row=row, column=3)
        ma_cell.value = avg if avg is not None else "NO DATA"
        ma_cell.font = DATA_FONT
        ma_cell.number_format = CURRENCY_FORMAT

        diff_cell = analysis.cell(row=row, column=4)
        diff_cell.value = f'=IF(OR(B{row}="",C{row}="",C{row}=0),"",B{row}-C{row})'
        diff_cell.font = DATA_FONT
        diff_cell.number_format = CURRENCY_FORMAT

        pct_cell = analysis.cell(row=row, column=5)
        pct_cell.value = f'=IF(OR(B{row}="",C{row}="",C{row}=0),"",(B{row}-C{row})/C{row})'
        pct_cell.font = DATA_FONT
        pct_cell.number_format = PERCENT_FORMAT

        updated += 1
        time.sleep(0.2)  # spread out ~80 sequential historical-candle calls

    wb.save(WORKBOOK_FILE)
    print(f"Updated 20-day MA for {updated} symbol(s) in {WORKBOOK_FILE}")


if __name__ == "__main__":
    main()

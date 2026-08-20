import datetime
import json
import os
import time
from collections import deque
from urllib.parse import quote

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

TOKEN_FILE = "token.json"
WORKBOOK_FILE = "watchlist.xlsx"
HISTORICAL_URL = "https://api.upstox.com/v3/historical-candle/{instrument_key}/days/1/{to_date}/{from_date}"

BACKTEST_YEARS = 3
WARMUP_CALENDAR_DAYS = 45  # buffer before backtest_start so the first day already has 20 trading days of history
TARGET_RETURN = 0.06
CHECKPOINT_RETURN = 0.03
MA_WINDOW = 20
MAX_HOLD_DAYS = 90  # force-exit at close if the +6% target hasn't hit within this many trading days
ENTRY_THRESHOLD_PCT = 0.02  # minimum fall required to take a signal (e.g. 0.02 = at least -2% below 20MA)

CAPITAL_TOTAL = 100000
MAX_SLOTS = 60
CAPITAL_PER_SLOT = CAPITAL_TOTAL / MAX_SLOTS

ARIAL = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name=ARIAL, size=10)
CURRENCY_FORMAT = "₹#,##0.00;-₹#,##0.00;-"
PERCENT_FORMAT = "0.00%;(0.00%);-"


def load_access_token():
    if not os.path.exists(TOKEN_FILE):
        raise SystemExit("No access token found - log in via /login first")
    with open(TOKEN_FILE) as f:
        return json.load(f)["access_token"]


def fetch_daily_candles(instrument_key, access_token, from_date, to_date):
    url = HISTORICAL_URL.format(
        instrument_key=quote(instrument_key, safe=""),
        to_date=to_date.isoformat(),
        from_date=from_date.isoformat(),
    )
    response = requests.get(
        url,
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
        },
    )
    response.raise_for_status()
    candles = response.json()["data"]["candles"]

    rows = []
    for c in candles:
        if c[5] <= 0:  # skip zero-volume / non-trading entries
            continue
        date = datetime.date.fromisoformat(c[0][:10])
        rows.append({"date": date, "open": c[1], "high": c[2], "low": c[3], "close": c[4]})
    rows.sort(key=lambda r: r["date"])
    return rows


def load_watchlist_symbols(wb):
    ws = wb["Watchlist"]
    symbols = []
    for row in range(5, ws.max_row + 1):
        symbol = ws.cell(row=row, column=1).value
        instrument_key = ws.cell(row=row, column=2).value
        if symbol and instrument_key and instrument_key != "NOT FOUND":
            symbols.append((str(symbol).strip().upper(), instrument_key))
    return symbols


def fetch_history(symbols, access_token, from_date, to_date):
    history = {}
    for i, (symbol, instrument_key) in enumerate(symbols, start=1):
        try:
            history[symbol] = fetch_daily_candles(instrument_key, access_token, from_date, to_date)
        except Exception as e:
            print(f"  [{i}/{len(symbols)}] skip {symbol}: {e}")
        else:
            print(f"  [{i}/{len(symbols)}] {symbol}: {len(history[symbol])} candles")
        time.sleep(0.2)
    return history


def days_to_reach(rows, entry_idx, entry_price, multiplier, last_idx):
    target = entry_price * multiplier
    for idx in range(entry_idx, last_idx + 1):
        if rows[idx]["high"] >= target:
            return idx - entry_idx
    return None


def run_backtest(history, backtest_start):
    symbol_dates = {s: [r["date"] for r in rows] for s, rows in history.items()}
    symbol_index = {s: {d: i for i, d in enumerate(dates)} for s, dates in symbol_dates.items()}
    all_dates = sorted({d for dates in symbol_dates.values() for d in dates if d >= backtest_start})

    open_positions = {}  # symbol -> {signal_date, entry_date, entry_price, entry_idx, compound_capital, signal_pct}
    free_slots = deque([CAPITAL_PER_SLOT] * MAX_SLOTS)
    fixed_realized_profit = 0.0
    compound_realized_profit = 0.0
    trades = []
    last_close = {}
    fixed_equity_curve = []
    compound_equity_curve = []

    for T in all_dates:
        for symbol, rows in history.items():
            idx = symbol_index[symbol].get(T)
            if idx is not None:
                last_close[symbol] = rows[idx]["close"]

        # 1. check exits for currently open positions (unlimited sells per day):
        #    target hit (+6%, checked against the day's High) or max holding period reached (forced exit at Close)
        for symbol in list(open_positions.keys()):
            idx = symbol_index[symbol].get(T)
            if idx is None:
                continue
            rows = history[symbol]
            row = rows[idx]
            pos = open_positions[symbol]
            if row["date"] < pos["entry_date"]:
                continue

            holding = idx - pos["entry_idx"]
            target_price = pos["entry_price"] * (1 + TARGET_RETURN)

            if row["high"] >= target_price:
                exit_price = target_price
                actual_return = TARGET_RETURN
                exit_reason = "TARGET"
            elif holding >= MAX_HOLD_DAYS:
                exit_price = row["close"]
                actual_return = (exit_price - pos["entry_price"]) / pos["entry_price"]
                exit_reason = "MAX HOLD"
            else:
                continue

            fixed_pnl = CAPITAL_PER_SLOT * actual_return
            compound_proceeds = pos["compound_capital"] * (1 + actual_return)
            compound_pnl = compound_proceeds - pos["compound_capital"]

            days_3pct = days_to_reach(rows, pos["entry_idx"], pos["entry_price"], 1 + CHECKPOINT_RETURN, idx)

            trades.append({
                "symbol": symbol,
                "signal_date": pos["signal_date"],
                "signal_pct": pos["signal_pct"],
                "entry_date": pos["entry_date"],
                "entry_price": pos["entry_price"],
                "exit_date": row["date"],
                "exit_price": exit_price,
                "return_pct": actual_return,
                "days_to_6pct": idx - pos["entry_idx"] if exit_reason == "TARGET" else None,
                "days_to_3pct": days_3pct,
                "status": "CLOSED",
                "exit_reason": exit_reason,
                "fixed_capital": CAPITAL_PER_SLOT,
                "fixed_pnl": fixed_pnl,
                "compound_capital": pos["compound_capital"],
                "compound_pnl": compound_pnl,
                "last_price": None,
                "unrealized_pct": None,
            })

            fixed_realized_profit += fixed_pnl
            compound_realized_profit += compound_pnl
            free_slots.append(compound_proceeds)
            del open_positions[symbol]

        # 2. compute today's signal (highest-fallen stock, only if beyond the entry threshold)
        candidates = []
        for symbol, rows in history.items():
            idx = symbol_index[symbol].get(T)
            if idx is None or idx < MA_WINDOW - 1:
                continue
            window = rows[idx - (MA_WINDOW - 1): idx + 1]
            ma20 = sum(r["close"] for r in window) / MA_WINDOW
            if ma20 == 0:
                continue
            close_t = rows[idx]["close"]
            pct_change = (close_t - ma20) / ma20
            candidates.append((pct_change, symbol, idx))

        if candidates:
            candidates.sort(key=lambda c: c[0])
            best_pct, best_symbol, best_idx = candidates[0]

            # at most 1 buy a day, and only if a capital slot is free
            if best_pct < -ENTRY_THRESHOLD_PCT and best_symbol not in open_positions and free_slots:
                rows = history[best_symbol]
                entry_idx = best_idx + 1
                if entry_idx < len(rows):
                    entry_row = rows[entry_idx]
                    open_positions[best_symbol] = {
                        "signal_date": T,
                        "signal_pct": best_pct,
                        "entry_date": entry_row["date"],
                        "entry_price": entry_row["open"],
                        "entry_idx": entry_idx,
                        "compound_capital": free_slots.popleft(),
                    }

        # 3. daily mark-to-market equity, for max-drawdown tracking
        fixed_open_value = sum(
            CAPITAL_PER_SLOT * (last_close.get(sym, pos["entry_price"]) / pos["entry_price"])
            for sym, pos in open_positions.items()
        )
        fixed_cash = CAPITAL_TOTAL - len(open_positions) * CAPITAL_PER_SLOT + fixed_realized_profit
        fixed_equity_curve.append((T, fixed_cash + fixed_open_value))

        compound_open_value = sum(
            pos["compound_capital"] * (last_close.get(sym, pos["entry_price"]) / pos["entry_price"])
            for sym, pos in open_positions.items()
        )
        compound_equity_curve.append((T, sum(free_slots) + compound_open_value))

    # remaining open positions at the end of the backtest window - mark to market
    for symbol, pos in open_positions.items():
        rows = history[symbol]
        last_idx = len(rows) - 1
        last_row = rows[last_idx]
        unrealized_pct = (last_row["close"] - pos["entry_price"]) / pos["entry_price"]
        days_3pct = days_to_reach(rows, pos["entry_idx"], pos["entry_price"], 1 + CHECKPOINT_RETURN, last_idx)

        trades.append({
            "symbol": symbol,
            "signal_date": pos["signal_date"],
            "signal_pct": pos["signal_pct"],
            "entry_date": pos["entry_date"],
            "entry_price": pos["entry_price"],
            "exit_date": None,
            "exit_price": None,
            "return_pct": None,
            "days_to_6pct": None,
            "days_to_3pct": days_3pct,
            "status": "OPEN",
            "exit_reason": None,
            "fixed_capital": CAPITAL_PER_SLOT,
            "fixed_pnl": CAPITAL_PER_SLOT * unrealized_pct,
            "compound_capital": pos["compound_capital"],
            "compound_pnl": pos["compound_capital"] * unrealized_pct,
            "last_price": last_row["close"],
            "unrealized_pct": unrealized_pct,
        })

    trades.sort(key=lambda t: t["entry_date"])

    fixed_unrealized = sum(t["fixed_pnl"] for t in trades if t["status"] == "OPEN")
    compound_unrealized = sum(t["compound_pnl"] for t in trades if t["status"] == "OPEN")

    def max_drawdown(curve):
        if not curve:
            return 0.0
        peak = curve[0][1]
        worst = 0.0
        for _, value in curve:
            peak = max(peak, value)
            worst = max(worst, (peak - value) / peak) if peak > 0 else worst
        return worst

    summary = {
        "capital_total": CAPITAL_TOTAL,
        "slots": MAX_SLOTS,
        "capital_per_slot": CAPITAL_PER_SLOT,
        "max_hold_days": MAX_HOLD_DAYS,
        "entry_threshold_pct": ENTRY_THRESHOLD_PCT,
        "total_trades": len(trades),
        "closed_trades": sum(1 for t in trades if t["status"] == "CLOSED"),
        "open_trades": sum(1 for t in trades if t["status"] == "OPEN"),
        "target_exits": sum(1 for t in trades if t.get("exit_reason") == "TARGET"),
        "max_hold_exits": sum(1 for t in trades if t.get("exit_reason") == "MAX HOLD"),
        "fixed_realized_profit": fixed_realized_profit,
        "fixed_unrealized_pnl": fixed_unrealized,
        "fixed_ending_value": CAPITAL_TOTAL + fixed_realized_profit + fixed_unrealized,
        "fixed_max_drawdown": max_drawdown(fixed_equity_curve),
        "compound_realized_profit": compound_realized_profit,
        "compound_unrealized_pnl": compound_unrealized,
        "compound_ending_value": sum(free_slots) + sum(
            t["compound_capital"] * (1 + t["unrealized_pct"]) for t in trades if t["status"] == "OPEN"
        ),
        "compound_max_drawdown": max_drawdown(compound_equity_curve),
    }
    return trades, summary


def write_backtest_sheet(wb, trades, summary):
    if "Backtest" in wb.sheetnames:
        del wb["Backtest"]
    ws = wb.create_sheet("Backtest")

    ws.merge_cells("A1:Q1")
    ws["A1"] = "Backtest - Buy Most-Fallen Stock, Sell at +6% (₹1,00,000 / 60 slots)"
    ws["A1"].font = Font(name=ARIAL, size=14, bold=True)

    ws.merge_cells("A2:Q2")
    ws["A2"] = (
        "Entry: next day's Open after signal. Exit: TARGET (day's High touches entry x 1.06) or "
        f"MAX HOLD (forced exit at Close after {summary['max_hold_days']} trading days). "
        "Max 60 concurrent positions, at most 1 new buy/day, unlimited sells/day. "
        "OPEN rows are marked-to-market at the last available close."
    )
    ws["A2"].font = Font(name=ARIAL, size=9, italic=True)

    summary_rows = [
        ("Total Capital", summary["capital_total"], CURRENCY_FORMAT),
        ("Slots", summary["slots"], "#,##0"),
        ("Capital / Slot", summary["capital_per_slot"], CURRENCY_FORMAT),
        ("Max Hold Days", summary["max_hold_days"], "#,##0"),
        ("Entry Threshold", -summary["entry_threshold_pct"], PERCENT_FORMAT),
        ("Total Trades", summary["total_trades"], "#,##0"),
        ("Closed / Open", f'{summary["closed_trades"]} / {summary["open_trades"]}', None),
        ("Exits: Target / Max Hold", f'{summary["target_exits"]} / {summary["max_hold_exits"]}', None),
        ("Fixed - Realized Profit", summary["fixed_realized_profit"], CURRENCY_FORMAT),
        ("Fixed - Unrealized P&L", summary["fixed_unrealized_pnl"], CURRENCY_FORMAT),
        ("Fixed - Ending Value", summary["fixed_ending_value"], CURRENCY_FORMAT),
        ("Fixed - Max Drawdown", summary["fixed_max_drawdown"], PERCENT_FORMAT),
        ("Compounding - Realized Profit", summary["compound_realized_profit"], CURRENCY_FORMAT),
        ("Compounding - Unrealized P&L", summary["compound_unrealized_pnl"], CURRENCY_FORMAT),
        ("Compounding - Ending Value", summary["compound_ending_value"], CURRENCY_FORMAT),
        ("Compounding - Max Drawdown", summary["compound_max_drawdown"], PERCENT_FORMAT),
    ]
    SUMMARY_LABEL_COL = 21  # column U, clear of the trade table (columns A-R, i.e. 1-18)
    SUMMARY_VALUE_COL = 23  # column W
    for i, (label, value, fmt) in enumerate(summary_rows):
        row = 4 + i
        label_cell = ws.cell(row=row, column=SUMMARY_LABEL_COL, value=label)
        label_cell.font = Font(name=ARIAL, size=10, bold=True)
        value_cell = ws.cell(row=row, column=SUMMARY_VALUE_COL, value=value)
        value_cell.font = DATA_FONT
        if fmt:
            value_cell.number_format = fmt
    ws.column_dimensions["U"].width = 26
    ws.column_dimensions["W"].width = 16

    headers = [
        "Symbol", "Signal Date", "Signal %", "Entry Date", "Entry Price", "Exit Date",
        "Exit Price", "Return %", "Days to +6%", "Days to +3%", "Status", "Exit Reason",
        "Fixed Capital", "Fixed P&L", "Compound Capital Used", "Compound P&L",
        "Last Price (open)", "Unrealized %",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(trades):
        row = header_row + 1 + i
        values = [
            t["symbol"], t["signal_date"], t["signal_pct"], t["entry_date"], t["entry_price"], t["exit_date"],
            t["exit_price"], t["return_pct"], t["days_to_6pct"], t["days_to_3pct"], t["status"], t["exit_reason"],
            t["fixed_capital"], t["fixed_pnl"], t["compound_capital"], t["compound_pnl"],
            t["last_price"], t["unrealized_pct"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            if col in (2, 4, 6):
                cell.number_format = "yyyy-mm-dd"
            elif col in (5, 7, 13, 14, 15, 16, 17):
                cell.number_format = CURRENCY_FORMAT
            elif col in (3, 8, 18):
                cell.number_format = PERCENT_FORMAT

    widths = [14, 12, 10, 12, 11, 12, 11, 10, 11, 11, 9, 11, 12, 11, 16, 12, 14, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=col).column_letter].width = width

    ws.freeze_panes = "A5"


def main():
    access_token = load_access_token()
    wb = load_workbook(WORKBOOK_FILE)
    symbols = load_watchlist_symbols(wb)

    to_date = datetime.date.today()
    backtest_start = to_date - datetime.timedelta(days=BACKTEST_YEARS * 365)
    from_date = backtest_start - datetime.timedelta(days=WARMUP_CALENDAR_DAYS)

    print(f"Fetching {len(symbols)} symbols from {from_date} to {to_date} ...")
    history = fetch_history(symbols, access_token, from_date, to_date)

    print("Running backtest ...")
    trades, summary = run_backtest(history, backtest_start)

    print(
        f"Trades: {summary['total_trades']} total, {summary['closed_trades']} closed "
        f"({summary['target_exits']} target, {summary['max_hold_exits']} max-hold), "
        f"{summary['open_trades']} open"
    )
    print(f"Fixed model    - ending value: Rs {summary['fixed_ending_value']:,.2f}, max drawdown: {summary['fixed_max_drawdown']:.2%}")
    print(f"Compounding    - ending value: Rs {summary['compound_ending_value']:,.2f}, max drawdown: {summary['compound_max_drawdown']:.2%}")

    signal_pcts = sorted(t["signal_pct"] for t in trades)
    if signal_pcts:
        n = len(signal_pcts)
        print(f"Signal %% at entry (n={n}): min {signal_pcts[0]:.2%}, median {signal_pcts[n // 2]:.2%}, max {signal_pcts[-1]:.2%}")
        for p in (10, 25, 50, 75, 90):
            idx = min(n - 1, int(n * p / 100))
            print(f"  p{p}: {signal_pcts[idx]:.2%}")

    write_backtest_sheet(wb, trades, summary)
    wb.save(WORKBOOK_FILE)
    print(f"Saved Backtest sheet to {WORKBOOK_FILE}")


if __name__ == "__main__":
    main()
